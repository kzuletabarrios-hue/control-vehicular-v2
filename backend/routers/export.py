# backend/routers/export.py
import io
from datetime import date
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db
from routers.auth import require_permiso

router = APIRouter()

NAVY = "0F2440"
AMBER = "F59E0B"


def _header_style(ws, headers: list[str]):
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def _stream(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/flota")
def export_flota(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("flota", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT placa, conductor, n_pallets, n_contenedores,
               cant_volumen_externo, muelle_cargue,
               fecha, hora_salida_muelle,
               ultima_tienda_visitada, protocolo, temperatura,
               sello, sello_entrada,
               fecha_salida, hora_salida_cedi,
               fecha_llegada, hora_llegada,
               observacion
        FROM flota_propia
        WHERE {' AND '.join(where)}
        ORDER BY fecha DESC, created_at DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Flota Propia"

    headers = [
        "Placa", "Conductor", "Pallets", "Contenedores",
        "Vol. Externo", "Muelle",
        "Fecha Registro", "H. Sal. Muelle",
        "Última Tienda", "Protocolo", "Temperatura",
        "Sello", "Sello Entrada",
        "Fecha Salida", "H. Salida CEDI",
        "Fecha Llegada", "H. Llegada", "Observación",
    ]
    _header_style(ws, headers)

    for row in rows:
        ws.append(list(row))

    _autowidth(ws)
    fname = f"flota_propia_{date.today()}.xlsx"
    return _stream(wb, fname)


@router.get("/proveedores")
def export_proveedores(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("proveedores", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT placa_vehiculo, nombre_conductor, tipo_vehiculo, empresa,
               muelle_descargue, carga_compartida,
               fecha, hora_ingreso,
               fecha_salida, hora_salida,
               actividad_a_desarrollar, dependencia_autoriza, fecha_pago_arl, observaciones
        FROM proveedores
        WHERE {' AND '.join(where)}
        ORDER BY fecha DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"

    headers = [
        "Placa", "Conductor", "Tipo Vehículo", "Empresa",
        "Muelle Descargue", "Carga Compartida",
        "Fecha Ingreso", "H. Ingreso",
        "Fecha Salida", "H. Salida",
        "Actividad", "Dependencia Autoriza", "Pago ARL", "Observaciones",
    ]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"proveedores_{date.today()}.xlsx")


@router.get("/control-acceso")
def export_control_acceso(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("control_acceso", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("ca.fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("ca.fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT ca.cedula, ca.nombre, ca.contratista,
               ca.fecha, ca.hora_ingreso,
               ca.fecha_salida, ca.hora_salida,
               ca.observaciones, b.estado AS estado_bd
        FROM control_acceso ca
        LEFT JOIN bd_control_acceso b ON ca.cedula = b.cedula
        WHERE {' AND '.join(where)}
        ORDER BY ca.fecha DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Control Acceso"

    headers = ["Cédula", "Nombre", "Contratista", "Fecha Ingreso", "H. Ingreso", "Fecha Salida", "H. Salida", "Observaciones", "Estado BD"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"control_acceso_{date.today()}.xlsx")


@router.get("/visitantes")
def export_visitantes(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("visitantes", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT nombre, cedula, empresa,
               fecha, hora_ingreso,
               fecha_salida, hora_salida, observaciones
        FROM visitantes
        WHERE {' AND '.join(where)}
        ORDER BY fecha DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Visitantes"

    headers = ["Nombre", "Cédula", "Empresa", "Fecha Ingreso", "H. Ingreso", "Fecha Salida", "H. Salida", "Observaciones"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"visitantes_{date.today()}.xlsx")


@router.get("/conductores")
def export_conductores(
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("flota", "export")),
):
    rows = db.execute(text("""
        SELECT codigo, conductor, n_cedula, celular, tipo, activo, created_at
        FROM conductores
        ORDER BY conductor ASC
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Conductores"

    headers = ["Código", "Nombre", "Cédula", "Celular", "Tipo", "Activo", "Creado"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"conductores_{date.today()}.xlsx")


@router.get("/bd-tiendas")
def export_bd_tiendas(
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("maestros", "read")),
):
    rows = db.execute(text("""
        SELECT codigo, name, direccion
        FROM distribucion
        ORDER BY codigo ASC NULLS LAST, name ASC
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tiendas"

    headers = ["Código", "Nombre", "Dirección"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"tiendas_{date.today()}.xlsx")


@router.get("/bd-vehiculos")
def export_bd_vehiculos(
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("maestros", "read")),
):
    rows = db.execute(text("""
        SELECT placa, marca, modelo, color, anio, tipo, capacidad, activo
        FROM vehiculos
        ORDER BY placa ASC
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    headers = ["Placa", "Marca", "Modelo", "Color", "Año", "Tipo", "Capacidad", "Activo"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"vehiculos_{date.today()}.xlsx")


@router.get("/bd-empleados")
def export_bd_empleados(
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("maestros", "read")),
):
    rows = db.execute(text("""
        SELECT cedula, nombre, contratista, estado
        FROM bd_control_acceso
        ORDER BY nombre ASC
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"

    headers = ["Cédula", "Nombre", "Contratista", "Estado"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"empleados_{date.today()}.xlsx")


@router.get("/bd-proveedores")
def export_bd_proveedores(
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("maestros", "read")),
):
    rows = db.execute(text("""
        SELECT nombre, nit, contacto, celular, activo, created_at
        FROM bd_proveedores
        ORDER BY nombre ASC
    """)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "BD Proveedores"

    headers = ["Nombre", "NIT", "Contacto", "Celular", "Activo", "Creado"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"bd_proveedores_{date.today()}.xlsx")


@router.get("/visita-vehicular")
def export_visita_vehicular(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("visita_vehicular", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT placa, conductor, motivo_visita,
               fecha, hora_ingreso,
               fecha_salida, hora_salida, observaciones
        FROM visita_vehicular
        WHERE {' AND '.join(where)}
        ORDER BY fecha DESC, created_at DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Visita Vehicular"

    headers = ["Placa", "Conductor", "Motivo Visita",
               "Fecha Ingreso", "H. Ingreso",
               "Fecha Salida", "H. Salida", "Observaciones"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"visita_vehicular_{date.today()}.xlsx")


@router.get("/sustancias")
def export_sustancias(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("control_acceso", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT fecha, descripcion, cantidad, responsable,
               fecha_salida, hora_salida, observaciones
        FROM sustancias
        WHERE {' AND '.join(where)}
        ORDER BY fecha DESC, created_at DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sustancias"

    headers = ["Fecha Ingreso", "Descripción", "Cantidad", "Responsable",
               "Fecha Salida", "H. Salida", "Observaciones"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"sustancias_{date.today()}.xlsx")


@router.get("/herramientas")
def export_herramientas(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("control_acceso", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT fecha, descripcion, cantidad, responsable,
               fecha_salida, hora_salida, observaciones
        FROM herramientas
        WHERE {' AND '.join(where)}
        ORDER BY fecha DESC, created_at DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Herramientas"

    headers = ["Fecha Ingreso", "Descripción", "Cantidad", "Responsable",
               "Fecha Salida", "H. Salida", "Observaciones"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"herramientas_{date.today()}.xlsx")


@router.get("/novedades")
def export_novedades(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("flota", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("n.fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("n.fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT u.nombre AS usuario,
               n.fecha, n.hora, n.modulo_origen, n.categoria,
               n.descripcion, n.estado
        FROM novedades n
        JOIN usuarios u ON u.id = n.usuario_id
        WHERE {' AND '.join(where)}
        ORDER BY n.fecha DESC, n.hora DESC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Novedades"

    headers = ["Usuario", "Fecha", "H. Registro", "Módulo", "Categoría", "Descripción", "Estado"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"novedades_{date.today()}.xlsx")


@router.get("/apoyos")
def export_apoyos(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("flota", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("a.fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("a.fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT u.nombre AS recorredor,
               a.fecha, a.motivo, a.tipo,
               a.hora_llegada, a.hora_salida
        FROM apoyos_operativos a
        JOIN usuarios u ON u.id = a.recorredor_id
        WHERE {' AND '.join(where)}
        ORDER BY a.fecha DESC, a.hora_llegada ASC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Apoyos Operativos"

    headers = ["Recorredor", "Fecha", "Motivo", "Tipo", "H. Llegada", "H. Salida"]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"apoyos_operativos_{date.today()}.xlsx")


@router.get("/rondas")
def export_rondas(
    fecha_desde: str = Query(default=None),
    fecha_hasta: str = Query(default=None),
    db: Session = Depends(get_db),
    _: dict = Depends(require_permiso("flota", "export")),
):
    where = ["1=1"]
    params = {}
    if fecha_desde:
        where.append("r.fecha >= :desde")
        params["desde"] = fecha_desde
    if fecha_hasta:
        where.append("r.fecha <= :hasta")
        params["hasta"] = fecha_hasta

    rows = db.execute(text(f"""
        SELECT u.nombre AS recorredor,
               r.fecha, rc.turno, rc.numero_ronda,
               p.nombre AS punto, p.orden AS orden_punto,
               r.hora_marcacion, r.estado, r.observacion
        FROM rondas r
        JOIN rondas_ciclos rc ON rc.id = r.ciclo_id
        JOIN puntos_ronda p   ON p.id  = r.punto_id
        JOIN usuarios u       ON u.id  = r.recorredor_id
        WHERE {' AND '.join(where)}
        ORDER BY r.fecha DESC, rc.numero_ronda ASC, p.orden ASC
    """), params).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rondas"

    headers = [
        "Recorredor", "Fecha", "Turno", "N° Ronda",
        "Punto", "Orden", "H. Marcación", "Estado", "Observación",
    ]
    _header_style(ws, headers)
    for row in rows:
        ws.append(list(row))
    _autowidth(ws)
    return _stream(wb, f"rondas_{date.today()}.xlsx")
